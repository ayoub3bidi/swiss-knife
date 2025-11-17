#!/usr/bin/env python3

import sys
import argparse
import json
import csv
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelReportGenerator:
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.current_row = 1
        
        # Styles
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        
        self.title_font = Font(bold=True, size=14)
        self.title_alignment = Alignment(horizontal="center")
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_workbook(self, sheet_name: str = "Report"):
        """Create new workbook."""
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name
        self.current_row = 1
    
    def add_title(self, title: str):
        """Add report title."""
        self.worksheet.merge_cells(f'A{self.current_row}:E{self.current_row}')
        cell = self.worksheet[f'A{self.current_row}']
        cell.value = title
        cell.font = self.title_font
        cell.alignment = self.title_alignment
        self.current_row += 2
    
    def add_metadata(self, metadata: Dict[str, str]):
        """Add report metadata."""
        for key, value in metadata.items():
            self.worksheet[f'A{self.current_row}'] = key
            self.worksheet[f'B{self.current_row}'] = value
            self.worksheet[f'A{self.current_row}'].font = Font(bold=True)
            self.current_row += 1
        self.current_row += 1
    
    def add_table(self, data: List[Dict], headers: Optional[List[str]] = None,
                 auto_filter: bool = True, freeze_panes: bool = True):
        """Add data table with headers."""
        if not data:
            return
        
        # Determine headers
        if headers is None:
            headers = list(data[0].keys())
        
        start_row = self.current_row
        
        # Add headers
        for col_idx, header in enumerate(headers, start=1):
            cell = self.worksheet.cell(row=self.current_row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        self.current_row += 1
        
        # Add data
        for row_data in data:
            for col_idx, header in enumerate(headers, start=1):
                cell = self.worksheet.cell(row=self.current_row, column=col_idx)
                value = row_data.get(header, '')
                
                # Handle numeric values
                if isinstance(value, (int, float)):
                    cell.value = value
                    if isinstance(value, float):
                        cell.number_format = '#,##0.00'
                else:
                    cell.value = str(value)
                
                cell.border = self.border
            
            self.current_row += 1
        
        # Auto-filter
        if auto_filter:
            self.worksheet.auto_filter.ref = f'A{start_row}:{get_column_letter(len(headers))}{self.current_row-1}'
        
        # Freeze panes
        if freeze_panes:
            self.worksheet.freeze_panes = f'A{start_row+1}'
        
        # Auto-size columns
        for col_idx, header in enumerate(headers, start=1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(header))
            for row in range(start_row + 1, self.current_row):
                cell_value = self.worksheet.cell(row=row, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            self.worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        self.current_row += 1
    
    def add_summary(self, data: List[Dict], numeric_columns: List[str]):
        """Add summary row with totals/averages."""
        if not data or not numeric_columns:
            return
        
        start_col = 1
        self.worksheet.cell(row=self.current_row, column=start_col).value = "SUMMARY"
        self.worksheet.cell(row=self.current_row, column=start_col).font = Font(bold=True)
        
        headers = list(data[0].keys())
        
        for col_name in numeric_columns:
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                
                # Calculate sum
                data_start = self.current_row - len(data)
                data_end = self.current_row - 1
                
                cell = self.worksheet.cell(row=self.current_row, column=col_idx)
                cell.value = f'=SUM({get_column_letter(col_idx)}{data_start}:{get_column_letter(col_idx)}{data_end})'
                cell.font = Font(bold=True)
                cell.number_format = '#,##0.00'
        
        self.current_row += 2
    
    def add_chart(self, chart_type: str, data_range: str, title: str,
                 position: str = None):
        """Add chart to worksheet."""
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'line':
            chart = LineChart()
        elif chart_type == 'pie':
            chart = PieChart()
        else:
            return
        
        chart.title = title
        chart.style = 10
        
        # Parse data range
        # Format: "A1:B10" or similar
        parts = data_range.split(':')
        if len(parts) == 2:
            data = Reference(self.worksheet, min_col=1, min_row=1,
                           max_col=5, max_row=self.current_row-1)
            chart.add_data(data, titles_from_data=True)
        
        # Position
        if position is None:
            position = f'A{self.current_row + 2}'
        
        self.worksheet.add_chart(chart, position)
        self.current_row += 15
    
    def load_csv(self, filepath: Path) -> List[Dict]:
        """Load data from CSV."""
        data = []
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Convert numeric strings
                processed = {}
                for key, value in row.items():
                    try:
                        if '.' in value:
                            processed[key] = float(value)
                        else:
                            processed[key] = int(value)
                    except (ValueError, AttributeError):
                        processed[key] = value
                data.append(processed)
        return data
    
    def load_json(self, filepath: Path) -> List[Dict]:
        """Load data from JSON."""
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        if isinstance(data, list):
            return data
        elif isinstance(data, dict):
            return [data]
        else:
            return []
    
    def generate_report(self, config: Dict, output_path: Path):
        """Generate report from configuration."""
        self.create_workbook(config.get('sheet_name', 'Report'))
        
        # Title
        if 'title' in config:
            self.add_title(config['title'])
        
        # Metadata
        if 'metadata' in config:
            metadata = config['metadata']
            metadata['Generated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.add_metadata(metadata)
        
        # Data
        if 'data' in config:
            data = config['data']
        elif 'data_file' in config:
            filepath = Path(config['data_file'])
            if filepath.suffix == '.csv':
                data = self.load_csv(filepath)
            elif filepath.suffix == '.json':
                data = self.load_json(filepath)
            else:
                data = []
        else:
            data = []
        
        if data:
            headers = config.get('headers')
            self.add_table(
                data,
                headers=headers,
                auto_filter=config.get('auto_filter', True),
                freeze_panes=config.get('freeze_panes', True)
            )
            
            # Summary
            if 'summary_columns' in config:
                self.add_summary(data, config['summary_columns'])
            
            # Charts
            if 'charts' in config:
                for chart_config in config['charts']:
                    self.add_chart(
                        chart_config['type'],
                        chart_config.get('data_range', 'A1:B10'),
                        chart_config['title'],
                        chart_config.get('position')
                    )
        
        # Save
        self.workbook.save(output_path)
        print(f"✓ Report generated: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate formatted Excel reports from data",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate from CSV
  python excel_report_generator.py data.csv -o report.xlsx
  
  # Generate from JSON config
  python excel_report_generator.py --config report_config.json -o report.xlsx
  
  # Quick report with title
  python excel_report_generator.py data.csv -o report.xlsx --title "Sales Report Q4"

Config file format (JSON):
{
  "title": "Monthly Sales Report",
  "sheet_name": "Sales Data",
  "metadata": {
    "Department": "Sales",
    "Period": "January 2024"
  },
  "data_file": "sales.csv",
  "headers": ["Product", "Units", "Revenue"],
  "summary_columns": ["Units", "Revenue"],
  "auto_filter": true,
  "freeze_panes": true,
  "charts": [
    {
      "type": "bar",
      "title": "Sales by Product",
      "data_range": "A1:C10"
    }
  ]
}
        """
    )
    
    parser.add_argument('input', nargs='?', type=Path,
                       help='Input data file (CSV or JSON)')
    parser.add_argument('-o', '--output', type=Path, required=True,
                       help='Output Excel file')
    
    parser.add_argument('--config', type=Path,
                       help='Configuration file (JSON)')
    parser.add_argument('--title', type=str,
                       help='Report title')
    parser.add_argument('--sheet-name', type=str, default='Report',
                       help='Worksheet name')
    
    args = parser.parse_args()
    
    if not HAS_OPENPYXL:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)
    
    try:
        generator = ExcelReportGenerator()
        
        if args.config:
            # Config-based generation
            with open(args.config, 'r') as f:
                config = json.load(f)
            generator.generate_report(config, args.output)
        
        elif args.input:
            # Simple generation from data file
            if args.input.suffix == '.csv':
                data = generator.load_csv(args.input)
            elif args.input.suffix == '.json':
                data = generator.load_json(args.input)
            else:
                print(f"Unsupported format: {args.input.suffix}")
                sys.exit(1)
            
            config = {
                'title': args.title or f'Report - {args.input.stem}',
                'sheet_name': args.sheet_name,
                'data': data
            }
            
            generator.generate_report(config, args.output)
        
        else:
            parser.error("Either input file or --config required")
        
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
